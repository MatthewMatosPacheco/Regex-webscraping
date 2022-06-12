from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook(r'C:\Users\matth\OneDrive\Documents\Grades.xlsx')
ws =wb.active
print(wb.sheetnames)
wb.create_sheet('test')
print(wb.sheetnames)# if you dont dave wont update
ws.title ="Data"
ws.append(["tim", "is","great"])# adds it horizontally on excel first sheet
ws.append(["tim", "is","great"])
ws.append(["tim", "is","great"])# every other line
ws.append(["end"])

for row in range (1,11):
    for col in range(1, 5):
        char= get_column_letter(col)
        # if you want to dothis manually char = chr(65 + col) #chr gives int of character A= 65 just because(uppercase)
        print(ws[char + str(row)].value)

ws.merged_cells("A1:D2")
ws.insert_rows(7)# insert empty row after 7
ws.insert_cols(7)
ws.delete_cols(7)

#shifting 2 cols
ws.move_range("C1:D11"' rows=2, col=2')
